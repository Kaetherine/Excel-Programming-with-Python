import barcode
from barcode.writer import ImageWriter
from bs4 import BeautifulSoup as bs4
from openpyxl import Workbook, load_workbook
import requests

URL = 'https://www.brunevollholz.de/'
KATALOG = []

def scrape_by_woodtype(wood_url):
    url = URL + wood_url

    homepage = requests.get(url).text

    soup = bs4(homepage, 'html.parser')

    tags = soup.find_all('td', class_ = 'col_1')

    count = -1
    for art in tags:
        count += 1

    object_count = int(count/ 4)

    for value in range(0, object_count+1):
        if art.get_text() != '':
            KATALOG.append([])

    i = 0
    for art in tags:
        if count <= 3:
            KATALOG[i].append(art.get_text())
            count += 1
        else:
            count = 0
            i += 1

    # for item in KATALOG:
    #     if item == []:
    #         KATALOG.remove(item)
    #         KATALOG.pop()
    #     else: pass
    # print(KATALOG,'\n')

WB = load_workbook('produktkatalog.xlsx')

def write_by_wood(sheet_title):

    ws = WB.create_sheet(sheet_title)
    for value in KATALOG:
        ws.append(value)

scrape_by_woodtype('saerge-eichenholz.html')
write_by_wood('Saerge Eichenholz')
KATALOG = []

scrape_by_woodtype('saerge-kiefernholz.html')
write_by_wood('Saerge Kiefernholz')
KATALOG = []

scrape_by_woodtype('saerge-andere-hoelzer.html')
write_by_wood('Saerge andere Hoelzer')
KATALOG = [] 

scrape_by_woodtype('saerge-kreativ-modelle.html')
write_by_wood('Saerge kreativ Modelle')

KATALOG = [] 
scrape_by_woodtype('saerge-exklusiv-modelle.html')
write_by_wood('Saerge exclusiv Modelle')

KATALOG = [] 
scrape_by_woodtype('saerge-sondergroessen.html')
write_by_wood('Saerge Sondergroessen')

WB.save('produktkatalog.xlsx')

